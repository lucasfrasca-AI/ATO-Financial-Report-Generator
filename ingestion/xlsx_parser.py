"""
xlsx_parser.py — openpyxl + pandas parser for Xero Excel exports.

Extracts:
  - Full text (concatenated cell values for redaction scanning)
  - Tables (one per sheet, as list of row dicts)

Xero exports typically have one financial statement per sheet.
Empty sheets are skipped.
"""

import logging
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


def parse_xlsx(path: str) -> dict:
    """
    Parse an Excel file (.xlsx/.xls) using openpyxl + pandas.

    Returns:
        {
            "name":   filename,
            "path":   absolute path,
            "type":   "xlsx",
            "text":   full text of all cells (for redaction),
            "tables": list of table dicts per sheet
        }
    """
    file_path = Path(path)

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    all_tables = []
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Read all cell values for text extraction (redaction scanning)
        cell_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cell_values.append(str(cell))

        if not cell_values:
            logger.debug("Sheet '%s' is empty — skipped", sheet_name)
            continue

        text_parts.append(f"[Sheet: {sheet_name}]")
        text_parts.extend(cell_values)

        # Use pandas for structured table extraction
        try:
            df = pd.read_excel(str(file_path), sheet_name=sheet_name, header=0, dtype=str)
            df = df.dropna(how="all").fillna("")
        except Exception as exc:
            logger.warning("pandas failed to parse sheet '%s': %s", sheet_name, exc)
            continue

        if df.empty:
            continue

        table = _df_to_table(df, sheet_name)
        if table:
            all_tables.append(table)

    full_text = "\n".join(text_parts).strip()

    if not full_text:
        raise ValueError(f"Excel file '{file_path.name}' appears to contain no data.")

    logger.debug(
        "XLSX parsed: %s — %d sheets with data, %d chars",
        file_path.name, len(all_tables), len(full_text),
    )

    return {
        "name": file_path.name,
        "path": str(file_path.resolve()),
        "type": "xlsx",
        "text": full_text,
        "tables": all_tables,
    }


def _df_to_table(df: pd.DataFrame, sheet_name: str) -> dict | None:
    """Convert a DataFrame to the standard table dict format."""
    if df.empty:
        return None

    headers = [str(col) for col in df.columns]
    rows = []
    for _, row in df.iterrows():
        row_dict = {str(k): str(v) for k, v in row.items()}
        # Skip rows that are entirely empty
        if all(v.strip() == "" for v in row_dict.values()):
            continue
        rows.append(row_dict)

    if not rows:
        return None

    return {
        "sheet": sheet_name,
        "headers": headers,
        "rows": rows,
    }
